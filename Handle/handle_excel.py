#coding=utf-8
import openpyxl
import os
import sys

# base_path=os.path.join(os.path.dirname(os.getcwd()), "Case\\New_mojie_test_case.xlsx")#获取当前路径的上一层路径
# op_ex=openpyxl.load_workbook(base_path)
# sheet_name=op_ex.sheetnames[0]
# wb = openpyxl.load_workbook(base_path)
# wr = op_ex.active
# wr.cell(1, 10, "008")
# wb.save(base_path)

"""
#1、加载excel
#2、加载Excel所有sheet的内容
#3、读取具体每一个表格的内容
#4、读取excle 的所有行数
#5、读取每一行的内容
#6、数据写入到一个单元格
#7、获取行号
#8、获取Excel的所有内容
"""

class Handle_Excel:
    base_path = os.path.join(os.path.dirname(os.getcwd()),'case\\New_mojie_test_case.xlsx')

    def load_excel(self):
        '''        1、初始化Excel，返回实例化对象        '''
        open_excel=openpyxl.load_workbook(self.base_path)
        return open_excel

    def get_sheet_data(self,index=None):
        '''
        2、加载Excel所有sheet的内容
        '''
        sheet_name=self.load_excel().sheetnames
        if index==None:
            index=0
        data=self.load_excel()[sheet_name[index]]

        return data

    def get_cell_data(self,row,cols,index=None):
        '''
        3、读取具体每一个表格的内容
        '''
        data=self.get_sheet_data(index).cell(row=row,column=cols).value
        return data

    def get_rows(self,index=None):
        # 4、读取excle 的所有行数
        if index == None:
            index = 0
        row=self.get_sheet_data(index).max_row
        return row

    def get_row_values(self,index=None):
        """ 5、根据传递的index(sheet表下标)值，选择读取数据读取某一行的所有内容"""
        data_list=[]
        for i in range(self.get_rows(index)-1):
            data_list.append(self.get_row_value(i+2,index))
        return data_list

    def get_row_value(self,row,index=None):
        '''
        6、根据传递的row（行号值）index(shett表下标)，选择读取数据读取某一行的内容
        '''
        data_list=[]
        for i in (self.get_sheet_data(index)[row]):
            data_list.append(i.value)
        return data_list

    def excel_write_data(self,row,cols,value):
        '''
        7、数据写入到一个单元格,row：行号，cols：列号
         '''
        wb = self.load_excel()
        wr = wb.active
        wr.cell(row,cols,value)
        wb.save(self.base_path)
    def liat_totuple(self,index):
        data=self.get_row_values(index)
        # print(len(data))
        for i in range(len(data)):
            data[i]=tuple(data[i])
        return data

if __name__ == '__main__':
    data=[[],[],[],[]]
    data[0]=tuple(data[0])
    print(data)
    ll=(1,2,3)
    print(ll[1])
    excel_case=Handle_Excel()
    # print (excel_case.get_sheet_data(0))
    # print (excel_case.get_sheet_data(2))
    print ('row',excel_case.get_rows(1))
    # print (excel_case.get_row_value(2,1))
    # print (excel_case.get_row_value(10,1))
    # print(type(excel_case.get_row_value(2, 1)))
    # print (excel_case.liat_totuple(1))
    print(type(excel_case.get_row_values(1)))
    # excel_case.get_row_value(2)
    # excel_case.get_row_value(2)
    # url=excel_case.get_cell_data(1,5)
    # print(url)
    # print (type(url))
    # excel_case.excel_write_data(1,8,"imooc_008")
    # data=Handle_Excel().get_cell_data(2,14,index=1)
    # data1=Handle_Excel().get_rows()
    # print(data1)
    # print(data)
